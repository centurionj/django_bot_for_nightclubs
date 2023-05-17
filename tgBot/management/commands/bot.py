from django.core.management.base import BaseCommand
from django.conf import settings

from telebot import TeleBot, types
from openpyxl import Workbook
from pathlib import Path

from tgBot.models import Day, Perfomace, Club, Order

bot = TeleBot(token=settings.TOKEN)

def send_info_of_day(id:int, message):
    perfomance = Perfomace.objects.filter(which_day_of_weak=id)
    for per in perfomance:
        s = f"\n{per.title}\n\n"
        try:
            media = per.preview.file
            if Path(f'{media}').suffix.lower() == '.mp4' or Path(f'{media}').suffix == '.mov':
                bot.send_document(message.chat.id, media, caption=s)
            else:
                bot.send_photo(message.chat.id, media, caption=s)
        except Exception as e:
            print(e)
            bot.send_message(message.chat.id, s)

def make_order():
    clubs = Club.objects.filter(qr_code=False)
    orders = Order.objects.order_by('club', 'date')

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "ФИО"
    ws['B1'] = "Клуб"
    ws['C1'] = "Дата"

    row_num = 2

    for order in orders:
        for club in clubs:
            if not club.qr_code and club.name.lower() == order.club.lower():
                row = [order.name_sername,
                       order.club,
                       str(order.date)[:11],
                       ]
                ws.append(row)
                row_num += 1

    for col in ws.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(cell.value)
            except Exception as e:
                print(e)
            adjusted_width = (max_len + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    wb.save('orders.xlsx')

def organisator(message):
    first = message
    new_message = first.text

    if new_message.lower() == 'admin':
        bot.send_message(message.chat.id, 'Отпраляю списки по клубам')
        make_order()
        doc1 = open('orders.xlsx', 'rb')
        bot.send_document(message.chat.id, doc1)
        menu(message)
    elif new_message == 'Назад' or new_message == 'назад':
        menu(message)
    else:
        bot.send_message(message.chat.id, 'Неверный логин')
        menu(message)

def folow(message):
    first = message
    new_message = first.text.lower()
    if new_message == 'назад':
        menu(message)
    else:
        try:
            ls = new_message.split()
            new_ls = [ls[0] + ' ' + ls[1], ls[2], ls[3]]
            my_list = [
                {'name_sername': f'{new_ls[0]}', 'club': f'{new_ls[1]}', 'date': f'{new_ls[2]}'},
            ]

            for item in my_list:
                order = Order(name_sername=item['name_sername'], club=item['club'], date=item['date'])
                order.save()
                bot.send_message(message.chat.id, 'Вы успешно добавлены в списки!')
                menu(message)
        except Exception as e:
            bot.send_message(message.chat.id, str(e))

def back_btn(message, txt):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    item = types.KeyboardButton('Назад')
    markup.add(item)
    bot.send_message(message.chat.id, txt, reply_markup=markup)


def menu(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    day1 = types.KeyboardButton('ПН')
    day2 = types.KeyboardButton('ВТ')
    day3 = types.KeyboardButton('СР')
    day4 = types.KeyboardButton('ЧТ')
    day5 = types.KeyboardButton('ПТ')
    day6 = types.KeyboardButton('СБ')
    day7 = types.KeyboardButton('ВС')
    registration = types.KeyboardButton('Записаться')
    workers = types.KeyboardButton('Организаторам')
    markup.add(day1, day2, day3, day4, day5, day6, day7, registration, workers)
    bot.send_message(message.chat.id, 'Выбери когда хочешь потусить', reply_markup=markup)


@bot.message_handler(commands=['start'])
def start(message):
    menu(message)

@bot.message_handler()
def bot_massage(message):
    if message.text.lower() == 'пн':
        send_info_of_day(1, message)

    elif message.text.lower() == 'вт':
        send_info_of_day(2, message)

    elif message.text.lower() == 'ср':
        send_info_of_day(3, message)

    elif message.text.lower() == 'чт':
        send_info_of_day(4, message)

    elif message.text.lower() == 'пт':
        send_info_of_day(5, message)

    elif message.text.lower() == 'сб':
        send_info_of_day(6, message)

    elif message.text.lower() == 'вс':
        send_info_of_day(7, message)

    elif message.text.lower() == 'записаться':
        txt = """Чтобы получить проходку в клубы: GIPSY, FANTASTIC, LOOKIN ROOMS вам нужно записаться здесь, в формате:
Фамилия Имя Клуб гггг-мм-дд

Всю информацию мы передадим в клубы, внесём вас в списки на проход. 

Также если вы хотите отдохнуть в
клубах PIPL, IZI, BIZI, ANIMA, выберите день недели, откройте описание мероприятия. Там вы найдёте ссылку на проходку. Регистрируйтесь, Qr-код- это ваш билет, его нужно будет показать на входе"""
        txt1 = 'Если попал сюда случайно нажми кнопку «Назад»'
        back_btn(message, txt)
        bot.send_message(message.chat.id, txt1)
        send = bot.send_message(message.chat.id, 'Будь внимателен при заполнении своих данных!')
        bot.register_next_step_handler(send, folow)

    elif message.text.lower() == 'организаторам':
        txt = 'Если попал сюда случайно нажми кнопку «Назад»'
        back_btn(message, txt)
        send = bot.send_message(message.chat.id, 'Введи логин')
        bot.register_next_step_handler(send, organisator)

    elif message.text.lower() == 'назад':
        menu(message)

    else:
        bot.send_message(message.chat.id, 'Такой команды я не знаю, выбери интересующий раздел из списка!')

class Command(BaseCommand):
    help = 'Тусовка бот'

    while True:
        try:
            bot.polling(none_stop=True)
        except:
            print('bolt')
            logging.error('error: {}'.format(sys.exc_info()[0]))


